"""
勤務分配表處理核心模組
將勤務表資料填入分配表模板
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from typing import Tuple, List, Optional
from io import BytesIO


def get_available_dates(duty_file) -> List[str]:
    """
    從勤務表取得所有可用的日期（工作表名稱）

    Args:
        duty_file: 上傳的勤務表檔案

    Returns:
        日期列表，如 ['0101', '0102', ...]
    """
    xls = pd.ExcelFile(duty_file)
    # 過濾掉「範本」等非日期的工作表
    dates = [name for name in xls.sheet_names if re.match(r'^\d{4}$', name)]
    # 按日期排序
    dates.sort()
    return dates


def format_number(value) -> str:
    """
    將數字轉換為兩位數字串（個位數補零）

    Args:
        value: 數字或字串

    Returns:
        格式化後的字串，如 "08"
    """
    if pd.isna(value):
        return ""

    # 處理純數字
    if isinstance(value, (int, float)):
        return f"{int(value):02d}"

    return str(value).strip()


def parse_duty_string(value) -> str:
    """
    將勤務表的人員編號格式轉換為分配表格式
    例如: "10.14.1" -> "10,14,01"

    Args:
        value: 勤務表中的值

    Returns:
        轉換後的字串
    """
    if pd.isna(value):
        return ""

    value_str = str(value).strip()
    if not value_str:
        return ""

    # 分割（可能用 . 或其他分隔符）
    # 處理像 "10.14.1" 或 "18.3" 這樣的格式
    parts = re.split(r'[.\s]+', value_str)

    # 轉換每個編號為兩位數
    formatted_parts = []
    for part in parts:
        part = part.strip()
        if part:
            try:
                num = int(float(part))
                formatted_parts.append(f"{num:02d}")
            except ValueError:
                formatted_parts.append(part)

    return ",".join(formatted_parts)


def parse_rescue_numbers(val1, val2) -> str:
    """
    合併兩個救護人員編號

    Args:
        val1: 第一救護
        val2: 第二救護

    Returns:
        合併後的字串，如 "08,20"
    """
    numbers = []

    for val in [val1, val2]:
        if pd.notna(val):
            try:
                num = int(float(val))
                numbers.append(f"{num:02d}")
            except ValueError:
                pass

    return ",".join(numbers)


def find_column_by_header(df, keywords: List[str], header_row: int = 2) -> Optional[int]:
    """
    根據標題關鍵字找到欄位索引

    Args:
        df: DataFrame
        keywords: 標題關鍵字列表（任一匹配即可）
        header_row: 標題所在的行號

    Returns:
        欄位索引，若找不到則返回 None
    """
    for col in range(df.shape[1]):
        cell_value = df.iloc[header_row, col]
        if pd.notna(cell_value):
            cell_str = str(cell_value).replace('\n', '').replace(' ', '')
            for keyword in keywords:
                if keyword in cell_str:
                    return col
    return None


def extract_duty_data(duty_file, date_sheet: str) -> dict:
    """
    從勤務表提取指定日期的勤務資料

    Args:
        duty_file: 勤務表檔案
        date_sheet: 工作表名稱（日期），如 "0120"

    Returns:
        包含所有勤務資料的字典
    """
    # 讀取指定工作表
    df = pd.read_excel(duty_file, sheet_name=date_sheet, header=None)

    # 根據標題名稱找到各欄位的位置
    col_duty = find_column_by_header(df, ['值班'])
    col_rescue = find_column_by_header(df, ['第一救護', '救護'])
    col_standby = find_column_by_header(df, ['備勤'])
    col_rest = find_column_by_header(df, ['休息'])

    # 時間對應：Row 4-27 對應 24 個時段
    # 08~09, 09~10, ..., 07~08
    time_slots = []
    for i in range(4, 28):
        # 讀取值班
        duty_val = ""
        if col_duty is not None:
            duty_val = format_number(df.iloc[i, col_duty])

        # 讀取救護（第一救護 + 下一欄的第二救護）
        rescue_val = ""
        if col_rescue is not None:
            rescue_val = parse_rescue_numbers(
                df.iloc[i, col_rescue],
                df.iloc[i, col_rescue + 1] if col_rescue + 1 < df.shape[1] else None
            )

        # 讀取備勤
        standby_val = ""
        if col_standby is not None:
            standby_val = parse_duty_string(df.iloc[i, col_standby])

        # 讀取休息
        rest_val = ""
        if col_rest is not None:
            rest_val = parse_duty_string(df.iloc[i, col_rest])

        time_slots.append({
            'time': df.iloc[i, 0] if pd.notna(df.iloc[i, 0]) else "",
            'duty': duty_val,
            'rescue': rescue_val,
            'standby': standby_val,
            'rest': rest_val,
        })

    # 提取輪休資訊 (Row 28)
    rotation_off = ""
    compensatory_off = ""
    if pd.notna(df.iloc[28, 3]):
        rotation_off = str(df.iloc[28, 3])  # 輪休人員
    if len(df.columns) > 9 and pd.notna(df.iloc[28, 9]):
        compensatory_off = str(df.iloc[28, 9])  # 補休人員

    # 提取備註 (Row 30)
    remarks = ""
    if pd.notna(df.iloc[30, 0]):
        remarks = str(df.iloc[30, 0])
        # 移除開頭的 "備註:" 或 "備註："
        remarks = re.sub(r'^備註[:：]\s*', '', remarks)
        # 將「。※」改成換行「\n※」
        remarks = remarks.replace('。※', '\n※')
        # 將全形冒號改成半形冒號
        remarks = remarks.replace('：', ':')
        # 移除結尾的句號
        remarks = remarks.rstrip('。')

    # 提取出動梯次 (Row 42)
    dispatch = ""
    if pd.notna(df.iloc[42, 2]):
        dispatch = str(df.iloc[42, 2])

        # 處理出動梯次格式：將括號內的數字補零
        # 例如 "16車(10.1)" -> "16車(10,01)"
        def format_dispatch_numbers(match):
            content = match.group(1)
            # 分割數字並補零
            parts = re.split(r'[.\s]+', content)
            formatted = []
            for p in parts:
                p = p.strip()
                if p:
                    try:
                        num = int(float(p))
                        formatted.append(f"{num:02d}")
                    except ValueError:
                        formatted.append(p)
            return "(" + ",".join(formatted) + ")"

        dispatch = re.sub(r'\(([^)]+)\)', format_dispatch_numbers, dispatch)

    # 提取車輛保養資訊 (Row 35-37)
    vehicle_maintenance_list = []
    for i in range(35, 38):
        if len(df.columns) > 22:
            vehicle_type = df.iloc[i, 19] if pd.notna(df.iloc[i, 19]) else ""
            vehicle_num = df.iloc[i, 20] if pd.notna(df.iloc[i, 20]) else ""
            maintainer = df.iloc[i, 22] if pd.notna(df.iloc[i, 22]) else ""
            if vehicle_type and vehicle_num and maintainer:
                vehicle_maintenance_list.append(f"({vehicle_type})保養車輛:{vehicle_num} 保養人:{maintainer}")
    vehicle_maintenance = "\n".join(vehicle_maintenance_list)

    # 固定附記文字（Row 44-47 的內容合併）
    fixed_note = "(號轄區消防查察、水源調查)及轄區搶救困難狹小巷道防火、防災、山難，獨居老人、防震、一氧化碳居家訪視、防颱宣導、學生寄宿舍、住宅火警警報器設備設置家戶宣導及AED及CPR教學、住宅老舊電線抽換電氣防範火災防火宣導、廟宇爆竹煙火使用有認可安全標示宣導及液化石油氣灌(分)裝場所、販賣場所取締逾期鋼瓶、超量儲存，及查稽可疑廢棄工寮及工廠、地下爆竹非法、製造、儲存場所取締及防溺宣導勤務"

    return {
        'time_slots': time_slots,
        'rotation_off': rotation_off,
        'compensatory_off': compensatory_off,
        'remarks': remarks,
        'dispatch': dispatch,
        'vehicle_maintenance': vehicle_maintenance,
        'fixed_note': fixed_note,
    }


def fill_distribution_table(template_file, duty_data: dict, date_str: str) -> BytesIO:
    """
    將勤務資料填入分配表模板

    Args:
        template_file: 分配表模板檔案
        duty_data: 從勤務表提取的資料
        date_str: 日期字串，如 "0120"

    Returns:
        填好資料的 Excel 檔案（BytesIO 格式）
    """
    # 載入模板（保留格式）
    wb = load_workbook(template_file)
    ws = wb.active

    # 填入時間欄位的資料 (Col 4-27 對應 24 個時段)
    # Row 3: 值班, Row 4: 救護, Row 10: 備勤, Row 11: 休息
    # 注意：openpyxl 的 row/col 是從 1 開始的

    time_slots = duty_data['time_slots']

    for i, slot in enumerate(time_slots):
        col = i + 5  # Col 4 在 openpyxl 中是第 5 欄（E欄）, 因為 pandas 是 0-indexed

        # Row 4: 值班 (pandas row 3 = openpyxl row 4)
        if slot['duty']:
            ws.cell(row=4, column=col, value=slot['duty'])

        # Row 5: 救護 (pandas row 4 = openpyxl row 5)
        if slot['rescue']:
            ws.cell(row=5, column=col, value=slot['rescue'])

        # Row 11: 備勤 (pandas row 10 = openpyxl row 11)
        if slot['standby']:
            ws.cell(row=11, column=col, value=slot['standby'])

        # Row 12: 休息 (pandas row 11 = openpyxl row 12)
        if slot['rest']:
            ws.cell(row=12, column=col, value=slot['rest'])

    # 填入出動梯次 (Row 17, Col 5)
    if duty_data['dispatch']:
        ws.cell(row=17, column=5, value=duty_data['dispatch'])

    # 填入附記 (Row 18, Col 5)
    # 格式: 備註 + 車輛保養（後面加空格）+ 固定附記文字
    remarks_parts = []

    # 1. 備註（晨間訓練、常年訓練等）
    if duty_data['remarks']:
        remarks_parts.append(duty_data['remarks'])

    # 2. 車輛保養資訊（後面加空格以符合原始格式）
    if duty_data['vehicle_maintenance']:
        # 在車輛保養後面加上空格（原始格式有留空格）
        remarks_parts.append(duty_data['vehicle_maintenance'] + "                      ")

    # 3. 固定附記文字
    if duty_data['fixed_note']:
        remarks_parts.append(duty_data['fixed_note'])

    remarks_text = "\n".join(remarks_parts)

    if remarks_text:
        ws.cell(row=18, column=5, value=remarks_text)

    # 儲存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def process_files(duty_file, template_file, date_sheet: str) -> Tuple[BytesIO, str]:
    """
    主處理函數：將勤務表資料填入分配表

    Args:
        duty_file: 勤務表檔案
        template_file: 分配表模板檔案
        date_sheet: 要處理的日期（工作表名稱）

    Returns:
        (填好的檔案, 建議的檔案名稱)
    """
    # 提取勤務資料
    duty_data = extract_duty_data(duty_file, date_sheet)

    # 填入分配表
    result_file = fill_distribution_table(template_file, duty_data, date_sheet)

    # 產生檔案名稱
    # 將 "0120" 轉換為 "20260120" 格式（假設是 115 年）
    month = date_sheet[:2]
    day = date_sheet[2:]
    filename = f"[2026{month}{day}] 屏二分隊勤務分配表.xlsx"

    return result_file, filename
